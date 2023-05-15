from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QComboBox, QLineEdit, QWidget, QTableWidget, QTableWidgetItem, QHBoxLayout, QSizePolicy, QSpacerItem, QVBoxLayout, QAbstractItemView
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QFont, QColor
from PyQt5 import QtGui
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


class MyWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Criando a tabela
        tabela_imc = QTableWidget(self)
        tabela_imc.setRowCount(7)
        tabela_imc.setColumnCount(2)
        item = QTableWidgetItem('Menor que 16,9')
        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(0, 0, item)
        tabela_imc.setHorizontalHeaderLabels(['IMC', 'Classificação'])
        # Primeira coluna
        tabela_imc.setItem(0, 0, QTableWidgetItem('Menor que 16,9'))
        tabela_imc.item(0, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(1, 0, QTableWidgetItem('Menor que 18,5'))
        tabela_imc.item(1, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(2, 0, QTableWidgetItem('18,5 a 24,9'))
        tabela_imc.item(2, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(3, 0, QTableWidgetItem('25 a 29,9'))
        tabela_imc.item(3, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(4, 0, QTableWidgetItem('30 a 34,9'))
        tabela_imc.item(4, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(5, 0, QTableWidgetItem('35 a 39,9'))
        tabela_imc.item(5, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(6, 0, QTableWidgetItem('Maior que 40'))
        tabela_imc.item(6, 0).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        # Segunda coluna
        tabela_imc.setItem(0, 1, QTableWidgetItem('Muito abaixo do peso'))
        tabela_imc.item(0, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(1, 1, QTableWidgetItem('Magreza'))
        tabela_imc.item(1, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(2, 1, QTableWidgetItem('Normal'))
        tabela_imc.item(2, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(3, 1, QTableWidgetItem('Sobrepeso'))
        tabela_imc.item(3, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(4, 1, QTableWidgetItem('Obesidade grau I'))
        tabela_imc.item(4, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(5, 1, QTableWidgetItem('Obesidade grau II (severa)'))
        tabela_imc.item(5, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        tabela_imc.setItem(6, 1, QTableWidgetItem('Obesidade Grau III (mórbida)'))
        tabela_imc.item(6, 1).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        tabela_imc.setColumnWidth(0, 200)
        tabela_imc.setColumnWidth(1, 200)


        # Define a cor desejada - Peso Abaixo
        azulc = tabela_imc.item(0, 0); azulc2 = tabela_imc.item(0, 1); yellow = tabela_imc.item(1, 0); yellow2 = tabela_imc.item(1, 1) 
        azulc.setBackground(QColor(135, 206, 250)); azulc2.setBackground(QColor(135, 206, 250)); yellow.setBackground(QColor(255, 255, 0)); yellow2.setBackground(QColor(255, 255, 0))
        #Peso normal e Sobre Peso
        green = tabela_imc.item(2, 0); green2 = tabela_imc.item(2, 1); sobre = tabela_imc.item(3, 0); sobre2 = tabela_imc.item(3, 1)
        green.setBackground(QColor(0, 148, 68)); green2.setBackground(QColor(0, 148, 68)); sobre.setBackground(QColor(255, 165, 0)); sobre2.setBackground(QColor(255, 165, 0))
        #Peso Obesidades
        obI = tabela_imc.item(4, 0); obI2 = tabela_imc.item(4, 1); obII = tabela_imc.item(5, 0); obII2 = tabela_imc.item(5, 1); 
        obI.setBackground(QColor(250,128,114)); obI2.setBackground(QColor(250,128,114)); obII.setBackground(QColor(245, 0, 0)); obII2.setBackground(QColor(245, 0, 0))
        obIII = tabela_imc.item(6, 0); obIII2 = tabela_imc.item(6, 1)
        obIII.setBackground(QColor(139, 0, 0)); obIII2.setBackground(QColor(139, 0, 0))

        # Criando o layout horizontal para adicionar a tabela ao centro
        hbox = QHBoxLayout()
        espacador_esquerda = QSpacerItem(290, 0, QSizePolicy.Fixed, QSizePolicy.Minimum)
        hbox.addItem(espacador_esquerda)
        hbox.addWidget(tabela_imc)
        tabela_imc.resizeColumnsToContents()
        tabela_imc.verticalHeader().setVisible(False)
        tabela_imc.setSelectionBehavior(QAbstractItemView.SelectItems)
        tabela_imc.setSelectionMode(QAbstractItemView.NoSelection)
        

        # Configurando o layout principal
        central_widget = QWidget(self)
        central_widget.setLayout(hbox)
        self.setCentralWidget(central_widget)


        # Configuração da janela principal
        self.setWindowTitle('Calculadora de IMC                                                                                                               coded by V1N1NH0')
        self.setGeometry(100, 100, 700, 335)

        # Criação dos elementos da interface
        idade_label = QLabel('Idade:', self); idade_label.setStyleSheet("color: white")
        sexo_label = QLabel('Sexo:', self); sexo_label.setStyleSheet("color: white")
        peso_label = QLabel('Peso (kg):', self); peso_label.setStyleSheet("color: white")
        altura_label = QLabel('Altura (m):', self); altura_label.setStyleSheet("color: white")
        self.resultado_label = QLabel('', self)
        self.resultimc_label = QLabel('', self)
        #self.resultado_label.setFixedWidth(3000)
        #self.resultado_label.setGeometry(50, 300, 600, 150)
        self.resultado_label.setGeometry(100, 600, 1000, 200)
        self.resultimc_label.setGeometry(100, 600, 1000, 200)
        # Criação dos elementos da interface
        self.imagem_label = QLabel(self)
        self.imagem_label.setGeometry(400, 50, 250, 250)
        #Imagem2 - Logo Etec
        self.imagem2_label = QLabel(self)
        self.imagem2_label.setGeometry(400, 50, 250, 250)
        #imagemOk - Simbolo Correto
        self.imagemOk_label = QLabel(self)
        self.imagemOk_label.setGeometry(400, 50, 250, 250)
        #imagem nivel saude mt baixo
        self.imagembaixo_label = QLabel(self)
        self.imagembaixo_label.setGeometry(400, 50, 250, 250)
        #imagem Alerta
        self.imagemalerta_label = QLabel(self)
        self.imagemalerta_label.setGeometry(400, 50, 250, 250)
        #Imagem Severa
        self.imagemsevera_label = QLabel(self)
        self.imagemsevera_label.setGeometry(400, 50, 250, 250)
        #Imagem Alert Red
        self.imagemalertred_label = QLabel(self)
        self.imagemalertred_label.setGeometry(400, 50, 250, 250)
        #Imagem mórbida
        self.imagemmorbida_label = QLabel(self)
        self.imagemmorbida_label.setGeometry(400, 50, 250, 250)


        # Criando um QLabel para exibir a imagem de fundo
        self.background = QLabel(self)
        self.background.lower()
        self.background.setGeometry(0, 0, 700, 335) #1334
        self.background.setPixmap(QtGui.QPixmap("image.png"))

        # Carrega a imagem
        imagem = QPixmap('vinin.png')
        self.imagem_label.setPixmap(imagem)
        imagem2 = QPixmap('imc25.png')
        self.imagem2_label.setPixmap(imagem2)

        self.idade_entry = QLineEdit(self)
        self.peso_entry = QLineEdit(self)
        self.altura_entry = QLineEdit(self)

        self.idade_entry.setMaxLength(2) # comprimento máximo de 2 caracteres
        self.idade_entry.setAlignment(Qt.AlignCenter) # alinhamento centralizado
        self.altura_entry.setAlignment(Qt.AlignCenter) # alinhamento centralizado
        self.altura_entry.setMaxLength(4) # comprimento máximo de 4 caracteres
        self.altura_entry.setPlaceholderText('0.00') # texto exibido antes da entrada de dados
        self.altura_entry.textChanged.connect(self.formatar_altura)
        
        self.peso_entry.setAlignment(Qt.AlignCenter) # alinhamento centralizado
        self.peso_entry.setMaxLength(5) # comprimento máximo de 5 caracteres
        self.peso_entry.setPlaceholderText('0.00') # texto exibido antes da entrada de dados
        self.peso_entry.textChanged.connect(self.formatar_peso)

        self.sexo_combobox = QComboBox(self)
        self.sexo_combobox.addItems(['Masculino', 'Feminino'])

        calcular_botao = QPushButton('Calcular', self)
        calcular_botao.clicked.connect(self.calcular_imc)

        #Deixar os itens arredondados:
        self.idade_entry.setStyleSheet('border-radius: 10px;')
        self.altura_entry.setStyleSheet('border-radius: 10px;')
        self.peso_entry.setStyleSheet('border-radius: 10px;')
        self.sexo_combobox.setStyleSheet('border-radius: 10px;')
        #calcular_botao.setStyleSheet('border-radius: 10px;')
        self.sexo_combobox.setStyleSheet('''
            QComboBox {
                border: 1px solid gray;
                border-radius: 10px;
                padding: 1px 18px 1px 3px;
                min-width: 6em;
                background-color: white;
            }

            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                width: 20px;
                border-left-width: 1px;
                border-left-color: gray;
                border-left-style: solid;
                border-top-right-radius: 10px;
                border-bottom-right-radius: 10px;
            }

            QComboBox::down-arrow {
                image: url('download.png');
                width: 12px;
                height: 12px;
            }
        ''')
        
        # Posicionamento dos elementos
        idade_label.move(50, 75)
        self.idade_entry.move(150, 75)

        sexo_label.move(50, 125)
        self.sexo_combobox.move(150, 125)

        peso_label.move(50, 175)
        self.peso_entry.move(150, 175)

        altura_label.move(50, 225)
        self.altura_entry.move(150, 225)

        calcular_botao.move(50, 275)

        self.resultimc_label.move(560, 5)
        self.resultado_label.move(310, 180)
        self.imagem_label.move(32, -85)
        self.imagem2_label.move(585, -77)
        self.imagemOk_label.move(590, 27)
        self.imagembaixo_label.move(590, 30)
        self.imagemalerta_label.move(585, 27)
        self.imagemsevera_label.move(585, 30)
        self.imagemalertred_label.move(575, 40)
        self.imagemmorbida_label.move(575, 40)

    def formatar_altura(self, text):
        text = text.replace(".", "")
        if len(text) >= 3:
            text = text[:-2] + "." + text[-2:]
        elif len(text) >= 2:
            text = text[:-1] + "." + text[-1:]
        self.altura_entry.setText(text)

    def formatar_peso(self, text):
        text = text.replace(".", "")
        if len(text) >= 5:
            text = text[:-1] + "." + text[:-1]
        elif len(text) >= 3:
            text = text[:-1] + "." + text[-1:]
        elif len(text) >= 2:
            text = text[:-1] + "." + text[-1:]
        self.peso_entry.setText(text)
    
    def calcular_imc(self):
        idade_texto = self.idade_entry.text().strip()
        sexo = self.sexo_combobox.currentText()
        peso_texto = self.peso_entry.text().strip()
        altura_texto = self.altura_entry.text().strip()

        # Verificar se todas as informações necessárias foram fornecidas
        if not idade_texto or not peso_texto or not altura_texto:
            self.resultado_label.setText('Por favor, preencha todos os campos.')
            return

        # Verificar se os valores de idade, peso e altura são válidos
        if not idade_texto.isdigit() or not peso_texto.replace('.', '', 1).isdigit() or not altura_texto.replace('.', '', 1).isdigit():
            self.resultado_label.setText('Por favor, forneça valores numéricos válidos.')
            return

        idade = int(idade_texto)
        peso = float(peso_texto)
        altura = float(altura_texto)

        imc = peso / (altura ** 2)
        peso_minimo = altura ** 2 * 18.5
        peso_maximo = altura ** 2 * 24.9 
        
        #limpar images

        if self.imagemOk_label.pixmap() is not None:
            self.imagemOk_label.clear()
        elif self.imagembaixo_label.pixmap() is not None:
            self.imagembaixo_label.clear()
        elif self.imagemalerta_label.pixmap() is not None:
            self.imagemalerta_label.clear()
        elif self.imagemsevera_label.pixmap() is not None:
            self.imagemsevera_label.clear()
        elif self.imagemalertred_label.pixmap() is not None:
            self.imagemalertred_label.clear()
        elif self.imagemmorbida_label.pixmap() is not None:
            self.imagemmorbida_label.clear()

        if imc < 16.9:
            nivel_saude = "Muito Abaixo do peso"; imagembaixo = QPixmap('mtbaixo.png'); self.imagembaixo_label.setPixmap(imagembaixo)
            cor = QColor(0, 0, 0) # azul claro
        elif imc < 18.5:
            nivel_saude = "Abaixo do peso"; imagemalerta = QPixmap('alert.png'); self.imagemalerta_label.setPixmap(imagemalerta) 
            cor = QColor(0, 0, 0) # amarelo claro 
        elif imc < 25:
            nivel_saude = "Peso normal (Parabéns)" ; imagemOk = QPixmap('imageok.png'); self.imagemOk_label.setPixmap(imagemOk)
            cor = QColor(0, 148, 68) # verde
        elif imc < 30:
            nivel_saude = "Sobrepeso"; imagemalerta = QPixmap('alert.png'); self.imagemalerta_label.setPixmap(imagemalerta)
            cor = QColor(0, 0, 0) # laranja claro
        elif imc < 35:
            nivel_saude = "Obesidade Grau I"; imagemalertred = QPixmap('alertred.png'); self.imagemalertred_label.setPixmap(imagemalertred) 
            cor = QColor(255, 140, 0) # laranja escuro 
        elif imc < 40:
            nivel_saude = "Obesidade Grau II (severa)"; imagemsevera = QPixmap('severa.jpg'); self.imagemsevera_label.setPixmap(imagemsevera)
            cor = QColor(255, 0, 0) # vermelho claro
        elif imc >= 40:
            nivel_saude = "Obesidade Grau III (mórbida)"; imagemmorbida = QPixmap('morbida.png'); self.imagemmorbida_label.setPixmap(imagemmorbida) 
            cor = QColor(139, 0, 0) # vermelho escuro


        self.resultado_label.setText(f"\nSeu nível de saúde é: {nivel_saude} \nPara manter o valor de IMC normal, \nseu peso pode variar entre {peso_minimo:.1f} e {peso_maximo:.1f} kg.")
        
        self.resultado_label.setStyleSheet("color: {}; font-size: 16px;".format(cor.name()))

        # IMC
        self.resultimc_label.setText(f"Seu IMC é: {imc:.2f}")
        
        self.resultimc_label.setStyleSheet("color: {}; font-size: 16px;".format(cor.name()))
        

        # Verifica se o arquivo já existe
        if os.path.exists("imc_data.xlsx"):
            # Carrega a planilha existente
            wb = openpyxl.load_workbook("imc_data.xlsx")
            ws = wb.active
        else:
            # Cria uma nova planilha
            wb = Workbook()
            ws = wb.active
            ws.title = "Dados IMC"
            headers = ["Idade", "Sexo", "Altura", "Peso", "IMC"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"] = header

        # Adiciona os dados na próxima linha vazia
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=idade)
        ws.cell(row=next_row, column=2, value=sexo)
        ws.cell(row=next_row, column=3, value=altura)
        ws.cell(row=next_row, column=4, value=peso)
        ws.cell(row=next_row, column=5, value=imc)

        # Salva a planilha
        wb.save("imc_data.xlsx")


# Inicia a aplicação
app = QApplication([])
window = MyWindow()
window.show()
app.exec_()

